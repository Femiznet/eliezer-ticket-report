import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from utils import log_error
from config import SAVE_TO_PATH

def _prepare_dataframe(data: list[dict], status: str) -> pd.DataFrame:
    df = pd.DataFrame(data)
    df["RequestCategory"] = df["RequestCategory"].fillna("NULL")
    
    date_cols = ["CreatedTime", "ClosedTime"]
    existing_date_cols = [col for col in date_cols if col in df.columns]
    
    df[existing_date_cols] = df[existing_date_cols].apply(
        lambda x: pd.to_datetime(x, utc=True, errors='coerce').dt.tz_localize(None)
    )
    
    sort_col = "CreatedTime" if status.lower() == "open" else "ClosedTime"
    if sort_col in df.columns:
        df = df.sort_values(by=sort_col, ascending=True)
        
    df[existing_date_cols] = df[existing_date_cols].apply(lambda x: x.dt.strftime("%d %b %Y"))
            
    return df

def _style_excel(file_path: str | Path):
    wb = load_workbook(file_path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        for i, col in enumerate(ws.columns, 1):
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 10), 50)
    wb.save(file_path)

def _export_tickets_to_excel(tickets: list[dict], assignee: str, status: str, base_folder: str = ".") -> Path:
    if not tickets:
        log_error("No data provided for export.")
        return
    
    df = _prepare_dataframe(tickets, status)
    
    output_dir = Path(base_folder) / assignee
    output_dir.mkdir(parents=True, exist_ok=True)
    file_path = output_dir / f"{status.lower()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    selected_cols = ["TicketNo", "Subject", "Address", "RequestCategory", "CreatedTime", "ClosedTime"]

    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="ALL", index=False)
        for category, group in df.groupby("RequestCategory"):
            group[selected_cols].to_excel(writer, sheet_name=str(category)[:31], index=False)
            
    _style_excel(file_path)

    return file_path

def convert_to_excel(data:dict, args):

    for each_owner in data:
        _export_tickets_to_excel(
            tickets=data[each_owner],
            assignee=each_owner,
            status=args.status,
            base_folder=SAVE_TO_PATH,
        )
